"""onedrive_mirror.py — write the customer book to OneDrive as an .xlsx.

One direction only: Postgres -> OneDrive. After each sync the current customer
book is rendered to Customer_Master.xlsx and uploaded to the ai@bharatsteels.in
drive, so the team always opens a fresh file. Nothing is ever read back from
OneDrive — the database is the source of truth, the file is a mirror.

Uses the same app-only Graph credentials pattern as the other BSG apps. The
Graph app registration needs Files.ReadWrite.All (application) with admin
consent, and GRAPH_DRIVE_USER must be a real, licensed mailbox/drive.

Skipped silently if GRAPH_* env vars aren't set, so the pipeline runs fine
without the mirror during early testing.
"""
from __future__ import annotations

import io
import logging
import os
import time

import httpx

log = logging.getLogger("mirror")

GRAPH = "https://graph.microsoft.com/v1.0"
TOKEN_URL = "https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"


def _cfg() -> dict | None:
    cfg = {
        "tenant": os.getenv("GRAPH_TENANT_ID") or os.getenv("MS_TENANT_ID", ""),
        "client": os.getenv("GRAPH_CLIENT_ID") or os.getenv("MS_CLIENT_ID", ""),
        "secret": os.getenv("GRAPH_CLIENT_SECRET") or os.getenv("MS_CLIENT_SECRET", ""),
        "drive_user": os.getenv("GRAPH_DRIVE_USER", "ai@bharatsteels.in"),
        "folder": os.getenv("GRAPH_MIRROR_FOLDER", "Enquiry Capture"),
        "filename": os.getenv("GRAPH_MIRROR_FILENAME", "Customer_Master.xlsx"),
    }
    if not (cfg["tenant"] and cfg["client"] and cfg["secret"]):
        return None
    return cfg


def _token(cfg: dict) -> str:
    r = httpx.post(
        TOKEN_URL.format(tenant=cfg["tenant"]),
        data={
            "client_id": cfg["client"],
            "client_secret": cfg["secret"],
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        },
        timeout=30,
    )
    r.raise_for_status()
    return r.json()["access_token"]


def _build_xlsx(rows: list[dict], summary: dict[str, int]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        "BP Code", "Company", "Contact Person", "Email", "Mobile",
        "Alt Emails", "Alt Mobiles", "GSTIN", "City", "Address", "Zip",
        "Source", "Times Seen", "Last Contact",
    ]
    ws.append(headers)

    for c in rows:
        ws.append([
            c.get("bp_code"), c.get("company_name"), c.get("contact_person"),
            c.get("email"), c.get("mobile"),
            ", ".join(c.get("email_alt") or []),
            ", ".join(c.get("mobile_alt") or []),
            c.get("gstin"), c.get("city"), c.get("address"), c.get("zip_code"),
            c.get("source"), c.get("times_seen"),
            c["last_contact_at"].strftime("%Y-%m-%d %H:%M")
                if c.get("last_contact_at") else "",
        ])

    # header style
    navy = PatternFill("solid", fgColor="002B5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = navy
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    widths = [12, 34, 22, 30, 13, 34, 24, 17, 14, 40, 8, 8, 11, 17]
    for col, w in zip("ABCDEFGHIJKLMN", widths):
        ws.column_dimensions[col].width = w

    # a small summary sheet so the reader can see it's live
    ws2 = wb.create_sheet("Book Status")
    ws2.append(["Metric", "Value"])
    ws2.append(["Total customers", len(rows)])
    ws2.append(["Have email", sum(1 for c in rows if c.get("email"))])
    ws2.append(["Have mobile", sum(1 for c in rows if c.get("mobile"))])
    ws2.append(["Missing email", sum(1 for c in rows if not c.get("email"))])
    ws2.append(["Missing mobile", sum(1 for c in rows if not c.get("mobile"))])
    ws2.append([])
    ws2.append(["Enriched in last 7 days (by field)", ""])
    for field, n in summary.items():
        ws2.append([f"  {field}", n])
    ws2.append([])
    ws2.append(["Generated", time.strftime("%Y-%m-%d %H:%M:%S")])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def mirror_to_onedrive(rows: list[dict], summary: dict[str, int]) -> bool:
    """Render + upload. Returns True on success, False if skipped, raises on hard error."""
    cfg = _cfg()
    if cfg is None:
        log.info("Mirror skipped — GRAPH_* / MS_* credentials not set")
        return False

    try:
        content = _build_xlsx(rows, summary)
    except ImportError:
        log.warning("Mirror skipped — openpyxl not installed")
        return False

    token = _token(cfg)
    # PUT to /drive/root:/<folder>/<file>:/content  (creates or replaces).
    path = f"{cfg['folder']}/{cfg['filename']}".replace("//", "/")
    url = (f"{GRAPH}/users/{cfg['drive_user']}/drive/root:/"
           f"{path}:/content")

    for attempt in range(3):
        r = httpx.put(
            url,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            content=content,
            timeout=120,
        )
        if r.status_code in (200, 201):
            item = r.json()
            log.info("Mirror uploaded: %s (%s rows) -> %s",
                     cfg["filename"], len(rows), item.get("webUrl", cfg["drive_user"]))
            return True
        if r.status_code == 404:
            raise RuntimeError(
                f"Mirror 404 — the drive for {cfg['drive_user']} was not found. "
                "Confirm it's a real, licensed mailbox with a OneDrive, and that "
                "the Graph app has Files.ReadWrite.All (application) with admin consent."
            )
        if r.status_code in (429, 503) or r.status_code >= 500:
            time.sleep(2 ** attempt)
            continue
        raise RuntimeError(f"Mirror upload failed [{r.status_code}]: {r.text[:300]}")
    raise RuntimeError("Mirror upload failed after retries")
