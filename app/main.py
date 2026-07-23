"""Optional HTTP surface — slots into the Metfraa Portal or runs standalone.

    uvicorn app.main:app --port 8080

Endpoints
    GET  /health              liveness + last sync status
    POST /sync                trigger a pass (this is what Cron hits)
    GET  /prospects           list, filterable
    GET  /prospects/{id}      one prospect + the messages that built it
    PATCH /prospects/{id}     human edit; edited fields become locked
    GET  /export.xlsx         download the whole table as Excel
"""
from __future__ import annotations

import io
import os
from typing import Any

from fastapi import FastAPI, Header, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from . import db
from .config import settings
from .sync import run_once

app = FastAPI(title="Enquiry Capture", version="1.0")

CRON_SECRET = os.getenv("CRON_SECRET", "")


def _auth(secret: str | None) -> None:
    if CRON_SECRET and secret != CRON_SECRET:
        raise HTTPException(401, "bad or missing x-cron-secret")


@app.get("/health")
def health() -> dict[str, Any]:
    with db.connect() as conn:
        row = conn.execute(
            "SELECT last_run_at, last_error, consecutive_errors "
            "FROM sync_state WHERE mailbox = %s", (settings.mailbox,)
        ).fetchone()
        counts = conn.execute(
            "SELECT status, count(*) AS n FROM prospects GROUP BY status"
        ).fetchall()
    return {
        "ok": True,
        "mailbox": settings.mailbox,
        "last_run_at": row["last_run_at"] if row else None,
        "last_error": row["last_error"] if row else None,
        "consecutive_errors": row["consecutive_errors"] if row else 0,
        "prospects_by_status": {c["status"]: c["n"] for c in counts},
    }


@app.post("/sync")
def sync(x_cron_secret: str | None = Header(default=None)) -> dict[str, Any]:
    _auth(x_cron_secret)
    stats = run_once()
    return stats.__dict__


@app.get("/prospects")
def list_prospects(
    status: str | None = Query(default=None),
    enquiry_type: str | None = Query(default=None),
    min_confidence: str | None = Query(default=None),
    limit: int = Query(default=100, le=1000),
    offset: int = 0,
) -> dict[str, Any]:
    where, params = ["1=1"], {}
    if status:
        where.append("status = %(status)s")
        params["status"] = status
    if enquiry_type:
        where.append("enquiry_type = %(etype)s")
        params["etype"] = enquiry_type
    if min_confidence:
        rank = {"low": 0, "medium": 1, "high": 2}
        if min_confidence not in rank:
            raise HTTPException(400, "min_confidence must be low|medium|high")
        keep = [k for k, v in rank.items() if v >= rank[min_confidence]]
        where.append("confidence = ANY(%(conf)s)")
        params["conf"] = keep

    sql = (
        "SELECT * FROM prospects WHERE " + " AND ".join(where) +
        " ORDER BY last_seen DESC LIMIT %(limit)s OFFSET %(offset)s"
    )
    params.update(limit=limit, offset=offset)

    with db.connect() as conn:
        rows = conn.execute(sql, params).fetchall()
        total = conn.execute(
            "SELECT count(*) AS n FROM prospects WHERE " + " AND ".join(where),
            params,
        ).fetchone()["n"]
    return {"total": total, "count": len(rows), "prospects": rows}


@app.get("/prospects/{pid}")
def get_prospect(pid: int) -> dict[str, Any]:
    with db.connect() as conn:
        row = conn.execute("SELECT * FROM prospects WHERE id = %s", (pid,)).fetchone()
        if not row:
            raise HTTPException(404, "not found")
        msgs = conn.execute(
            "SELECT id, received_at, subject, from_address, verdict, body_snippet "
            "FROM messages WHERE prospect_id = %s ORDER BY received_at DESC",
            (pid,),
        ).fetchall()
    return {"prospect": row, "messages": msgs}


class ProspectPatch(BaseModel):
    company_name: str | None = None
    contact_person: str | None = None
    mobile: str | None = None
    email: str | None = None
    gstin: str | None = None
    city: str | None = None
    enquiry_type: str | None = None
    status: str | None = None
    notes: str | None = None


@app.patch("/prospects/{pid}")
def patch_prospect(pid: int, patch: ProspectPatch) -> dict[str, Any]:
    """Human edit. Any field set here is added to locked_fields, so the
    importer will never quietly overwrite it on a later email."""
    changes = {k: v for k, v in patch.model_dump().items() if v is not None}
    if not changes:
        raise HTTPException(400, "nothing to update")

    lockable = [k for k in changes if k not in ("status", "notes")]
    set_sql = ", ".join(f"{k} = %({k})s" for k in changes)

    with db.connect() as conn:
        row = conn.execute(
            f"""UPDATE prospects
                   SET {set_sql},
                       locked_fields = ARRAY(
                           SELECT DISTINCT unnest(locked_fields || %(lock)s::text[])
                       ),
                       updated_at = now()
                 WHERE id = %(id)s
             RETURNING *""",  # noqa: S608 — keys are pydantic-constrained
            {**changes, "id": pid, "lock": lockable},
        ).fetchone()
    if not row:
        raise HTTPException(404, "not found")
    return row


@app.get("/export.xlsx")
def export_xlsx() -> StreamingResponse:
    """Excel is a *view* here, generated on demand — not the database."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "pip install openpyxl to enable export")

    with db.connect() as conn:
        rows = conn.execute(
            "SELECT company_name, contact_person, email, mobile, gstin, city, "
            "enquiry_type, confidence, status, times_seen, first_seen, last_seen "
            "FROM prospects ORDER BY last_seen DESC"
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    headers = [
        "Company", "Contact Person", "Email", "Mobile", "GSTIN", "City",
        "Type", "Confidence", "Status", "Times Seen", "First Seen", "Last Seen",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r["company_name"], r["contact_person"], r["email"], r["mobile"],
            r["gstin"], r["city"], r["enquiry_type"], r["confidence"],
            r["status"], r["times_seen"],
            r["first_seen"].strftime("%Y-%m-%d %H:%M") if r["first_seen"] else "",
            r["last_seen"].strftime("%Y-%m-%d %H:%M") if r["last_seen"] else "",
        ])

    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="002B5F")
    for col, width in zip("ABCDEFGHIJKL", [34, 22, 30, 14, 18, 16, 12, 12, 12, 11, 17, 17]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="prospects.xlsx"'},
    )
